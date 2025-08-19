import tkinter as tk
from tkinter import ttk, messagebox, filedialog, PhotoImage
import database
import sys
import os
import member_form
import reporting_window
import settings_window
import csv
import tempfile
import platform
import subprocess
from datetime import datetime
# NOTE: we *intentionally* do not import openpyxl at top-level.
# We'll import it lazily inside _show_import_dialog() if needed.


class MemberApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Dug Hill Rod & Gun Club Membership Database")
        self.root.geometry("1100x600")

        self.base_dir = os.path.dirname(os.path.abspath(__file__))

        # ----- Menubar -----
        menubar = tk.Menu(self.root)
        self.root.config(menu=menubar)

        # File menu
        file_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="File", menu=file_menu)
        file_menu.add_command(label="➕ Add Member", command=self.add_member)
        file_menu.add_command(label="✏️ Edit Selected", command=self.edit_selected)
        file_menu.add_command(label="❌ Delete Selected", command=self.delete_selected)
        file_menu.add_separator()
        file_menu.add_command(label="Import ⬆️", command=self._show_import_dialog)
        file_menu.add_command(label="Export ⬇️", command=self._show_export_dialog)
        file_menu.add_separator()
        file_menu.add_command(label="Print Current Tab 🖨", command=self._print_members)
        file_menu.add_separator()
        file_menu.add_command(label="Exit", command=self.root.quit)

        # Members menu
        members_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="Members", menu=members_menu)
        members_menu.add_command(label="Add Member", command=self.add_member)

        # Reports menu
        reports_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="Reports", menu=reports_menu)
        reports_menu.add_command(
            label="Dues Report",
            command=lambda: reporting_window.ReportingWindow(self.root)
        )
        # Recycle Bin
        menubar.add_command(label="Recycle Bin", command=self._show_recycle_bin)

        # Settings menu
        menubar.add_command(label="Settings", command=self.open_settings)
        
        # track recycle bin refresh function
        self.recycle_bin_refresh_fn = None

        # Load icons
        self.icons = {}
        try:
            self.icons["add"] = PhotoImage(file=os.path.join(self.base_dir, "icons", "add.png"))
            self.icons["edit"] = PhotoImage(file=os.path.join(self.base_dir, "icons", "edit.png"))
            self.icons["delete"] = PhotoImage(file=os.path.join(self.base_dir, "icons", "delete.png"))
            self.icons["recycle"] = PhotoImage(file=os.path.join(self.base_dir, "icons", "recycle.png"))
            self.icons["export"] = PhotoImage(file=os.path.join(self.base_dir, "icons", "export.png"))
            self.icons["import"] = PhotoImage(file=os.path.join(self.base_dir, "icons", "import.png"))
        except Exception:
            self.icons = {k: None for k in ["add", "edit", "delete", "recycle", "export", "import"]}

        # Style
        style = ttk.Style(self.root)
        style.theme_use("default")

        style.configure(
            "Custom.Vertical.TScrollbar",
            troughcolor="#f0f0f0",
            background="#d0d0d0",
            bordercolor="#a0a0a0",
            arrowcolor="black"
        )
        style.configure(
            "Custom.Horizontal.TScrollbar",
            troughcolor="#f0f0f0",
            background="#d0d0d0",
            bordercolor="#a0a0a0",
            arrowcolor="black"
        )

        self.member_types = [
            "All",
            "Probationary",
            "Associate",
            "Active",
            "Life",
            "Prospective",
            "Wait List",
            "Former"
        ]

        self.trees = {}
        self.all_members_data = []

        self._build_gui()
        self.load_data()

        # --- Keyboard shortcuts ---
        self.root.bind_all("<Control-p>", lambda e: self._print_members())   # Windows/Linux
        self.root.bind_all("<Command-p>", lambda e: self._print_members())   # macOS

    def _make_tree_with_scrollbars(self, parent, columns):
        container = ttk.Frame(parent)
        container.pack(fill="both", expand=True)

        container.grid_rowconfigure(0, weight=1)
        container.grid_columnconfigure(0, weight=1)

        tree = ttk.Treeview(container, columns=columns, show="headings", selectmode="extended")
        for col in columns:
            if col == "ID":
                tree.column(col, width=0, stretch=False)
                tree.heading(col, text="", anchor="center")
            else:
                tree.heading(col, text=col,
                             command=lambda _col=col, _tree=tree: self._sort_treeview(_tree, _col, False))
                tree.column(col, width=120, anchor="w")

        yscroll = ttk.Scrollbar(container, orient="vertical", style="Custom.Vertical.TScrollbar")
        xscroll = ttk.Scrollbar(parent, orient="horizontal", style="Custom.Horizontal.TScrollbar")
        yscroll.grid(row=0, column=1, sticky="ns")
        xscroll.pack(side="bottom", fill="x")

        tree.configure(yscrollcommand=yscroll.set, xscrollcommand=xscroll.set)
        yscroll.config(command=tree.yview)
        xscroll.config(command=tree.xview)

        tree.grid(row=0, column=0, sticky="nsew")
        return tree

    def _build_gui(self):
        # Toolbar
        toolbar = tk.Frame(self.root)
        toolbar.pack(side="top", fill="x")

        # Search bar + buttons
        search_frame = tk.Frame(toolbar)
        search_frame.pack(fill="x", padx=5, pady=3)

        #tk.Label(search_frame, text="Search:").pack(side="right")
        self.search_var = tk.StringVar()
        search_entry = ttk.Entry(search_frame, textvariable=self.search_var, width=40, justify="left")
        search_entry.pack(side="right", padx=5)
        search_entry.bind("<KeyRelease>", self._on_search)
        tk.Label(search_frame, text="Search:").pack(side="right")
        
        # Quick Buttons
        
        #ttk.Button(search_frame, text="🗑️ Recycle Bin", command=self._show_recycle_bin).pack(side="right", padx=2)
        #ttk.Button(search_frame, text="❌ Delete Selected", command=self.delete_selected).pack(side="right", padx=2)
        #ttk.Button(search_frame, text="✏️ Edit Selected", command=self.edit_selected).pack(side="right", padx=2)
        #ttk.Button(search_frame, text="➕ Add Member", command=self.add_member).pack(side="right", padx=2)

        # Notebook tabs
        self.notebook = ttk.Notebook(self.root)
        self.notebook.pack(fill="both", expand=True)

        columns = (
            "ID", "Badge Number", "Membership Type", "First Name", "Last Name",
            "Date of Birth", "Email Address", "Email Address 2", "Phone Number",
            "Address", "City", "State", "Zip Code", "Join Date", "Sponsor",
            "Card/Fob Internal Number", "Card/Fob External Number",
        )

        self.context_menu_main = tk.Menu(self.root, tearoff=0)
        self.context_menu_main.add_command(label="✏️ Edit", command=self.edit_selected)
        self.context_menu_main.add_command(label="❌ Delete", command=self.delete_selected)
        
        self.context_menu_blank = tk.Menu(self.root, tearoff=0)
        self.context_menu_blank.add_command(label="➕ Add Member", command=self.add_member)

        for mtype in self.member_types:
            frame = ttk.Frame(self.notebook)
            self.notebook.add(frame, text=mtype)

            tree = self._make_tree_with_scrollbars(frame, columns)
            tree.bind("<Double-1>", self._on_tree_double_click)
            tree.bind("<Button-3>", self._on_right_click_main)
            tree.bind("<Button-2>", self._on_right_click_main)
            tree.bind("<Control-Button-1>", self._on_right_click_main)
            tree.bind("<Button-1>", self._on_tree_click, add="+")


            self.trees[mtype] = tree
    def _on_tree_click(self, event):
        tree = event.widget
        row_id = tree.identify_row(event.y)
        if row_id == "":
            # Clicked on empty area → clear current selection
            tree.selection_remove(tree.selection())

    # ---------------- SORTING ---------------- #
    def _sort_treeview(self, tree, col, reverse):
        items = [(tree.set(k, col), k) for k in tree.get_children("")]
        try:
            if col == "Badge Number":
                items.sort(key=lambda t: int(t[0]) if str(t[0]).isdigit() else float("inf"), reverse=reverse)
            else:
                items.sort(key=lambda t: t[0].lower() if isinstance(t[0], str) else t[0], reverse=reverse)
        except Exception:
            items.sort(reverse=reverse)

        for index, (_, k) in enumerate(items):
            tree.move(k, "", index)
        tree.heading(col, command=lambda: self._sort_treeview(tree, col, not reverse))

    # ---------------- PRINT MEMBERS ---------------- #
    def _print_members(self):
        current_tab = self.notebook.tab(self.notebook.select(), "text")
        tree = self.trees[current_tab]

        items = tree.get_children()
        if not items:
            messagebox.showwarning("Print", "No members to print in this view.")
            return

        headings = [tree.heading(col)["text"] for col in tree["columns"]]
        rows = [headings]
        for item in items:
            rows.append(tree.item(item, "values"))

        col_widths = [max(len(str(row[i])) for row in rows) for i in range(len(headings))]
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        width_sum = sum(col_widths) + 3 * (len(col_widths) - 1)

        lines = [
            "Membership List".center(width_sum),
            f"Tab: {current_tab}".center(width_sum),
            f"Generated: {timestamp}".center(width_sum),
            ""
        ]

        for idx, row in enumerate(rows):
            line = "   ".join(str(row[i]).ljust(col_widths[i]) for i in range(len(row)))
            if idx == 0:
                lines.append(line)
                lines.append("-" * len(line))
            else:
                lines.append(line)
        lines.append("")
        lines.append("End of List".center(width_sum))

        report_text = "\n".join(lines)

        preview = tk.Toplevel(self.root)
        preview.title("Print Preview - Members")
        preview.geometry("800x600")

        text = tk.Text(preview, wrap="none")
        text.insert("1.0", report_text)
        text.configure(state="disabled")
        text.pack(fill="both", expand=True, padx=5, pady=5)

        yscroll = ttk.Scrollbar(preview, orient="vertical", command=text.yview)
        xscroll = ttk.Scrollbar(preview, orient="horizontal", command=text.xview)
        text.configure(yscrollcommand=yscroll.set, xscrollcommand=xscroll.set)
        yscroll.pack(side="right", fill="y")
        xscroll.pack(side="bottom", fill="x")

        btn_frame = tk.Frame(preview)
        btn_frame.pack(fill="x", pady=5)

        def do_print():
            tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".txt")
            tmp.write(report_text.encode("utf-8"))
            tmp.close()
            try:
                system = platform.system()
                if system == "Windows":
                    os.startfile(tmp.name, "print")
                elif system == "Darwin":
                    subprocess.run(["lp", tmp.name], check=False)
                else:
                    subprocess.run(["lpr", tmp.name], check=False)
            except Exception as e:
                messagebox.showerror("Print Error", f"Failed to print: {e}")

        ttk.Button(btn_frame, text="🖨 Print", command=do_print).pack(side="right", padx=5)
        ttk.Button(btn_frame, text="Close", command=preview.destroy).pack(side="right", padx=5)

    # ---------------- EXPORT ---------------- #
    def _show_export_dialog(self):
        dialog = tk.Toplevel(self.root)
        dialog.title("Export Members")
        dialog.geometry("300x300")

        tk.Label(dialog, text="Select Membership Types:").pack(anchor="w", padx=10, pady=5)

        vars = {}
        for mtype in self.member_types:
            var = tk.BooleanVar(value=(mtype == "All"))
            chk = tk.Checkbutton(dialog, text=mtype, variable=var)
            chk.pack(anchor="w", padx=15)
            vars[mtype] = var

        def do_export():
            selected_types = [mtype for mtype, var in vars.items() if var.get()]
            if not selected_types:
                messagebox.showwarning("No Selection", "Please select at least one membership type.")
                return

            now_str = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"members_{'_'.join(selected_types)}_{now_str}.csv"
            filepath = filedialog.asksaveasfilename(defaultextension=".csv", initialfile=filename,
                                                    filetypes=[("CSV files", "*.csv")])
            if not filepath:
                return

            try:
                with open(filepath, "w", newline="", encoding="utf-8") as f:
                    writer = csv.writer(f)
                    headers = (
                        "ID", "Badge Number", "Membership Type", "First Name", "Last Name",
                        "Date of Birth", "Email Address", "Email Address 2", "Phone Number",
                        "Address", "City", "State", "Zip Code", "Join Date", "Sponsor",
                        "Card/Fob Internal Number", "Card/Fob External Number"
                    )
                    writer.writerow(headers)

                    for row in self.all_members_data:
                        if "All" in selected_types or row[2] in selected_types:
                            writer.writerow(row)
            except Exception as e:
                messagebox.showerror("Export Failed", f"Could not write file:\n{e}")
                return

            messagebox.showinfo("Export Complete", f"Data exported to {filepath}")
            dialog.destroy()

        ttk.Button(dialog, text="Export", command=do_export).pack(pady=10)

    # ---------------- IMPORT ---------------- #
    def _show_import_dialog(self):
        filepath = filedialog.askopenfilename(
            title="Import Members",
            filetypes=[("CSV files", "*.csv"), ("Excel files", "*.xlsx")]
        )
        if not filepath:
            return

        imported_rows = []
        try:
            if filepath.endswith(".csv"):
                with open(filepath, newline="", encoding="utf-8") as f:
                    reader = csv.DictReader(f)
                    for row in reader:
                        imported_rows.append(row)
            elif filepath.endswith(".xlsx"):
                try:
                    import openpyxl  # lazy import
                except ImportError:
                    messagebox.showerror(
                        "Import Failed",
                        "Excel import requires the 'openpyxl' package.\n\nInstall it with:\n  pip install openpyxl"
                    )
                    return

                wb = openpyxl.load_workbook(filepath)
                sheet = wb.active
                headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    imported_rows.append(dict(zip(headers, row)))
        except Exception as e:
            messagebox.showerror("Import Failed", f"Error reading file: {e}")
            return

        count = 0
        for row in imported_rows:
            try:
                database.insert_member_from_dict(row)
                count += 1
            except Exception as e:
                print(f"Skipping row due to error: {e}")

        messagebox.showinfo("Import Complete", f"Imported {count} members.")
        self.load_data()

    # ---------------- DATA LOADING ---------------- #
    def load_data(self):
        """Reload all member data into every tree. Fully safe against app shutdown."""
        try:
            for tree in self.trees.values():
                tree.delete(*tree.get_children())
        except tk.TclError:
            return

        try:
            members = database.get_all_members()
        except Exception as e:
            try:
                messagebox.showerror("Database Error", f"Failed to load members: {e}")
            except tk.TclError:
                pass
            return

        self.all_members_data = []
        try:
            for m in members:
                data = (
                    str(m[0] or ""),   # ID
                    str(m[1] or ""),   # Badge Number
                    str(m[2] or ""),   # Membership Type
                    str(m[3] or ""),   # First Name
                    str(m[4] or ""),   # Last Name
                    str(m[5] or ""),   # Date of Birth
                    str(m[6] or ""),   # Email Address
                    str(m[13] or ""),  # Email Address 2
                    str(m[7] or ""),   # Phone Number
                    str(m[8] or ""),   # Address
                    str(m[9] or ""),   # City
                    str(m[10] or ""),  # State
                    str(m[11] or ""),  # Zip Code
                    str(m[12] or ""),  # Join Date
                    str(m[14] or ""),  # Sponsor
                    str(m[15] or ""),  # Card/Fob Internal Number
                    str(m[16] or ""),  # Card/Fob External Number
                )
                self.all_members_data.append(data)

                # Insert into "All"
                tree_all = self.trees.get("All")
                if tree_all:
                    tree_all.insert("", "end", values=data)

                # Insert into specific membership-type tab (if present)
                member_type = m[2]
                tree_type = self.trees.get(member_type)
                if tree_type:
                    tree_type.insert("", "end", values=data)
        except tk.TclError:
            return

    def _on_search(self, event=None):
        """Live filter across all tabs; safe against app shutdown."""
        query = (self.search_var.get() or "").lower()

        try:
            for tree in self.trees.values():
                tree.delete(*tree.get_children())

            filtered = (
                [row for row in self.all_members_data if any(query in str(val).lower() for val in row)]
                if query else self.all_members_data
            )

            # Refill per tab
            for data in filtered:
                # All
                t_all = self.trees.get("All")
                if t_all:
                    t_all.insert("", "end", values=data)

                # Specific type
                mt = data[2]
                t_type = self.trees.get(mt)
                if t_type:
                    t_type.insert("", "end", values=data)
        except tk.TclError:
            return

    # ---------------- RECYCLE BIN ---------------- #
    def _show_recycle_bin(self):
        win = tk.Toplevel(self.root)
        win.title("Recycle Bin")
        win.geometry("800x400")

        cols = (
            "ID", "Badge Number", "Membership Type", "First Name", "Last Name",
            "Date of Birth", "Email Address", "Email Address 2", "Phone Number",
            "Address", "City", "State", "Zip Code", "Join Date", "Sponsor",
            "Card/Fob Internal Number", "Card/Fob External Number", "Deleted At",
        )

        tree = self._make_tree_with_scrollbars(win, cols)

        def refresh_recycle_bin():
            try:
                for row in tree.get_children():
                    tree.delete(row)
                deleted_members = database.get_deleted_members()
                for m in deleted_members:
                    data = (
                        str(m[0] or ""), str(m[1] or ""), str(m[2] or ""), str(m[3] or ""),
                        str(m[4] or ""), str(m[5] or ""), str(m[6] or ""), str(m[13] or ""),
                        str(m[7] or ""), str(m[8] or ""), str(m[9] or ""), str(m[10] or ""),
                        str(m[11] or ""), str(m[12] or ""), str(m[14] or ""), str(m[15] or ""),
                        str(m[16] or ""), str(m[17] or "")
                    )
                    tree.insert("", "end", values=data)
            except tk.TclError:
                return

        refresh_recycle_bin()
        self.recycle_bin_refresh_fn = refresh_recycle_bin

        # Buttons
        ttk.Button(win, text="Restore Selected",
                   command=lambda: self.restore_selected(tree, refresh_recycle_bin)).pack(pady=5)
        ttk.Button(win, text="Delete Permanently",
                   command=lambda: self.permanent_delete_selected(tree, refresh_recycle_bin)).pack(pady=5)
        tree.bind("<Double-1>", lambda event, _tree=tree: self._on_double_click_deleted(event, _tree, refresh_recycle_bin))

        # --- Right-click menu for recycle bin ---
        menu_recycle = tk.Menu(win, tearoff=0)
        menu_recycle.add_command(label="Restore Selected",
                                 command=lambda: self.restore_selected(tree, refresh_recycle_bin))
        menu_recycle.add_command(label="Delete Permanently",
                                 command=lambda: self.permanent_delete_selected(tree, refresh_recycle_bin))

        def on_recycle_right_click(event, _tree=tree, _menu=menu_recycle):
            row_id = _tree.identify_row(event.y)
            if row_id:
                _tree.selection_set(row_id)
            try:
                _menu.tk_popup(event.x_root, event.y_root)
            finally:
                _menu.grab_release()

        tree.bind("<Button-3>", on_recycle_right_click)
        tree.bind("<Button-2>", on_recycle_right_click)
        tree.bind("<Control-Button-1>", on_recycle_right_click)

        def on_close():
            self.recycle_bin_refresh_fn = None
            win.destroy()

        win.protocol("WM_DELETE_WINDOW", on_close)

    # ---------------- MEMBER OPS ---------------- #
    def add_member(self):
        try:
            form = self._open_member_form()
            self.root.wait_window(form.top)
        except tk.TclError:
            return

    def edit_selected(self):
        try:
            current_tab = self.notebook.tab(self.notebook.select(), "text")
            tree = self.trees[current_tab]
            selected = tree.selection()
            if not selected:
                messagebox.showwarning("No selection", "Please select a member to edit.")
                return
            member_id = tree.item(selected[0])["values"][0]
            form = self._open_member_form(member_id)
            self.root.wait_window(form.top)
        except tk.TclError:
            return

    def delete_selected(self):
        try:
            current_tab = self.notebook.tab(self.notebook.select(), "text")
            tree = self.trees[current_tab]
            selected = tree.selection()
            if not selected:
                messagebox.showwarning("No selection", "Please select one or more members to delete.")
                return
            if not messagebox.askyesno("Confirm", "Are you sure you want to delete selected members?"):
                return

            for sel in selected:
                member_id = tree.item(sel)["values"][0]
                database.soft_delete_member_by_id(member_id)

            self.load_data()
            if self.recycle_bin_refresh_fn:
                self.recycle_bin_refresh_fn()
        except tk.TclError:
            return

    def restore_selected(self, tree, refresh_recycle_bin):
        try:
            selected = tree.selection()
            if not selected:
                messagebox.showwarning("No selection", "Please select a member to restore.")
                return
            for sel in selected:
                member_id = tree.item(sel)["values"][0]
                database.restore_member(member_id)
            self.load_data()
            refresh_recycle_bin()
        except tk.TclError:
            return

    def permanent_delete_selected(self, tree, refresh_recycle_bin):
        try:
            selected = tree.selection()
            if not selected:
                messagebox.showwarning("No selection", "Please select a member to delete permanently.")
                return
            if not messagebox.askyesno("Permanent Delete", "This cannot be undone. Delete permanently?"):
                return
            for sel in selected:
                member_id = tree.item(sel)["values"][0]
                database.permanent_delete_member(member_id)
            refresh_recycle_bin()
        except tk.TclError:
            return

    def _on_tree_double_click(self, event):
        try:
            tree = event.widget
            selected = tree.selection()
            if not selected:
                return
            member_id = tree.item(selected[0])["values"][0]
            form = self._open_member_form(member_id)
            self.root.wait_window(form.top)
        except tk.TclError:
            return

    def _on_double_click_deleted(self, event, tree, refresh_recycle_bin):
        try:
            selected = tree.selection()
            if not selected:
                return
            member_id = tree.item(selected[0])["values"][0]
            if messagebox.askyesno("Restore Member", "Do you want to restore this member?"):
                database.restore_member(member_id)
                self.load_data()
                refresh_recycle_bin()
        except tk.TclError:
            return

    # --- Right-click handler for main trees ---
    def _on_right_click_main(self, event):
        try:
            tree = event.widget
            row_id = tree.identify_row(event.y)

            if row_id:  # right-clicked on a row
                if row_id not in tree.selection():
                    tree.selection_set(row_id)
                menu = self.context_menu_main
            else:       # right-clicked on blank space
                menu = self.context_menu_blank

            try:
                menu.tk_popup(event.x_root, event.y_root)
            finally:
                menu.grab_release()
        except tk.TclError:
            return


    def _open_member_form(self, member_id=None):
        # Define callback for after save (accepts id and optional type)
        def on_save_callback(saved_id, saved_type=None):
            self.load_data()
            # Find and select the row for this member_id
            for mtype, tree in self.trees.items():
                for row in tree.get_children():
                    values = tree.item(row, "values")
                    if str(values[0]) == str(saved_id):  # first col = ID
                        tree.selection_set(row)
                        tree.focus(row)
                        tree.see(row)
                        return

        form = member_form.MemberForm(self.root, member_id, on_save_callback=on_save_callback)
        return form

    def open_settings(self):
        settings_window.SettingsWindow(self.root)


if __name__ == "__main__":
    root = tk.Tk()
    app = MemberApp(root)
    root.mainloop()

